import json
import os
import shutil
import time
import zipfile
from statistics import mean

import openpyxl

from Sewage import settings
from conf import my_setting
from lib.common import list_split


def ex_container(monitor_objs):
    """
    导出采用容器法的监测点数据
    :param monitor_objs: 监测点对象集
    :return: time_str + '.zip': 压缩包名
    """

    time_str = time.time()
    time_str = str(time_str)  # 时间戳作为压缩包和临时文件夹名
    output_path = os.path.join(my_setting.export_folder, time_str)
    shutil.rmtree(my_setting.export_folder)
    if not os.path.exists(my_setting.export_folder):
        os.mkdir(my_setting.export_folder)
    if not os.path.exists(output_path):
        os.mkdir(output_path)
    for monitor_count, monitor in enumerate(monitor_objs, 1):
        output_path_per_monitor = os.path.join(output_path, monitor.name)  # 存放每个监测点相应文件的临时路径
        output_path_for_photo_per_monitor = os.path.join(output_path_per_monitor, '照片')  # 存放每个监测点相应照片的临时路径
        output_path_for_excel_per_monitor = os.path.join(output_path_per_monitor, '成果')  # 存放每个监测点相应Excel文件的临时路径
        if not os.path.exists(output_path_per_monitor):
            os.mkdir(output_path_per_monitor)
        if not os.path.exists(output_path_for_photo_per_monitor):
            os.mkdir(output_path_for_photo_per_monitor)
        if not os.path.exists(output_path_for_excel_per_monitor):
            os.mkdir(output_path_for_excel_per_monitor)
        shutil.copy(os.path.join(my_setting.excel_folder, '小区监测统计表.xlsx'),
                    output_path_for_excel_per_monitor)  # 监测点对应统计表excel表格
        shutil.copy(os.path.join(my_setting.excel_folder, '流量表.xlsx'),
                    output_path_for_excel_per_monitor)  # 监测点对应流量表excel表格
        shutil.copy(os.path.join(my_setting.excel_folder, '监测点信息.xlsx'), output_path_per_monitor)  # 监测点对应监测点信息excel表格

        """20190417增加总体的小区监测统计表"""
        if not os.path.exists(os.path.join(output_path, '小区监测统计表.xlsx')):
            shutil.copy(os.path.join(my_setting.excel_folder, '小区监测统计表.xlsx'),
                        output_path)  # 监测点对应统计表excel表格
        from openpyxl.styles import Border, Side

        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        # 监测点信息表
        info_excel = openpyxl.load_workbook(os.path.join(output_path_per_monitor, '监测点信息.xlsx'))
        monitor_sheet = info_excel['监测点']
        monitor_sheet['A{}'.format(2)] = monitor.id
        monitor_sheet['B{}'.format(2)] = monitor.name
        monitor_sheet['C{}'.format(2)] = monitor.geophysical_point
        monitor_sheet['D{}'.format(2)] = monitor.people
        if monitor.work_function == 0:
            monitor_sheet['E{}'.format(2)] = '容器法'
        elif monitor.work_function == 1:
            monitor_sheet['E{}'.format(2)] = '流速法（圆管）'
        elif monitor.work_function == 2:
            monitor_sheet['E{}'.format(2)] = '流速法（方渠）'
        elif monitor.work_function == 3:
            monitor_sheet['E{}'.format(2)] = '仪器法'
        else:
            monitor_sheet['E{}'.format(2)] = '无法监测'
        monitor_sheet['F{}'.format(2)] = monitor.not_monitor_reason
        # 外景照
        exterior_photo_lst = json.loads(monitor.exterior_photo)
        exterior_photo_lst = list_split(exterior_photo_lst)
        for exterior_photo in exterior_photo_lst:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, exterior_photo), output_path_for_photo_per_monitor)
        # 水流照
        water_flow_photo_lst = json.loads(monitor.water_flow_photo)
        water_flow_photo_lst = list_split(water_flow_photo_lst)
        for water_photo in water_flow_photo_lst:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, water_photo), output_path_for_photo_per_monitor)
        # 工作照
        work_photo_lst = json.loads(monitor.work_photo)
        work_photo_lst = list_split(work_photo_lst)
        for work_photo in work_photo_lst:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, work_photo), output_path_for_photo_per_monitor)

        # 样品
        sample_sheet = info_excel['样品']
        sample_sheet_max_row = sample_sheet.max_row
        samples = monitor.sample.all().order_by('sample_date', 'sample_time')
        static_excel = openpyxl.load_workbook(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))
        static_sample = static_excel['水质']

        """20190417增加总体的小区监测统计表"""
        static_excel_total = openpyxl.load_workbook(os.path.join(output_path, '小区监测统计表.xlsx'))
        static_sample_total = static_excel_total['水质']
        static_sample_total_max_row = static_sample_total.max_row

        for sample_count, sample in enumerate(samples, 1):
            sample_sheet['A{}'.format(sample_sheet_max_row + sample_count)] = sample.id
            sample_sheet['B{}'.format(sample_sheet_max_row + sample_count)] = sample.monitor_point.id
            sample_sheet['C{}'.format(sample_sheet_max_row + sample_count)] = sample.people
            sample_sheet['D{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_date
            sample_sheet['E{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_time
            sample_sheet['F{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_number
            sample_sheet['G{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_color
            sample_sheet['H{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_odor
            sample_sheet['I{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_turbidity
            sample_sheet['J{}'.format(sample_sheet_max_row + sample_count)] = sample.monitor_task
            sample_sheet['K{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_count
            sample_sheet['L{}'.format(sample_sheet_max_row + sample_count)] = sample.SS
            sample_sheet['M{}'.format(sample_sheet_max_row + sample_count)] = sample.NH3_N
            sample_sheet['N{}'.format(sample_sheet_max_row + sample_count)] = sample.TP
            sample_sheet['O{}'.format(sample_sheet_max_row + sample_count)] = sample.TN
            sample_sheet['P{}'.format(sample_sheet_max_row + sample_count)] = sample.COD
            sample_sheet['Q{}'.format(sample_sheet_max_row + sample_count)] = sample.BOD
            sample_sheet['R{}'.format(sample_sheet_max_row + sample_count)] = sample.AIS
            sample_sheet['S{}'.format(sample_sheet_max_row + sample_count)] = sample.AFVO
            sample_sheet['T{}'.format(sample_sheet_max_row + sample_count)] = sample.DO
            sample_sheet['U{}'.format(sample_sheet_max_row + sample_count)] = sample.FLOW
            sample_sheet['V{}'.format(sample_sheet_max_row + sample_count)] = sample.CR
            sample_sheet['W{}'.format(sample_sheet_max_row + sample_count)] = sample.ORP
            sample_sheet['X{}'.format(sample_sheet_max_row + sample_count)] = sample.SinkableS
            sample_sheet['Y{}'.format(sample_sheet_max_row + sample_count)] = sample.Sulfide
            sample_sheet['Z{}'.format(sample_sheet_max_row + sample_count)] = sample.Cyanide
            # 样品照
            sample_photo_lst = json.loads(sample.sample_photo)
            sample_photo_lst = list_split(sample_photo_lst)
            for sample_photo in sample_photo_lst:
                shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, sample_photo), output_path_for_photo_per_monitor)
            # 打开样品统计表
            static_sample['A{}'.format(4 + sample_count)] = sample_count  # 序号
            static_sample['B{}'.format(4 + sample_count)] = monitor.name  # 采样点位
            static_sample['E{}'.format(4 + sample_count)] = monitor.geophysical_point  # 物探点号
            static_sample['H{}'.format(4 + sample_count)] = sample.sample_date  # 采样日期
            static_sample['I{}'.format(4 + sample_count)] = sample.sample_time  # 采样时间段
            static_sample['J{}'.format(4 + sample_count)] = sample.sample_number  # 样品编号
            static_sample[
                'Q{}'.format(
                    4 + sample_count)] = sample.sample_color + '、' + sample.sample_odor + '、' + sample.sample_turbidity  # 样品状态
            static_sample['R{}'.format(4 + sample_count)] = sample.monitor_task  # 监测项目
            static_sample['S{}'.format(4 + sample_count)] = sample.sample_count  # 样品数量
            static_sample['T{}'.format(4 + sample_count)] = sample.people  # 采样人

            """20190417增加总体的小区监测统计表"""
            # 打开样品统计表
            static_sample_total['A{}'.format(static_sample_total_max_row + sample_count)] = static_sample_total[
                                                                                                'A{}'.format(
                                                                                                    static_sample_total_max_row + sample_count)].row - 4  # 序号
            static_sample_total['B{}'.format(static_sample_total_max_row + sample_count)] = monitor.name  # 采样点位
            static_sample_total[
                'E{}'.format(static_sample_total_max_row + sample_count)] = monitor.geophysical_point  # 物探点号
            static_sample_total['H{}'.format(static_sample_total_max_row + sample_count)] = sample.sample_date  # 采样日期
            static_sample_total['I{}'.format(static_sample_total_max_row + sample_count)] = sample.sample_time  # 采样时间段
            static_sample_total['J{}'.format(static_sample_total_max_row + sample_count)] = sample.sample_number  # 样品编号
            static_sample_total[
                'Q{}'.format(
                    static_sample_total_max_row + sample_count)] = sample.sample_color + '、' + sample.sample_odor + '、' + sample.sample_turbidity  # 样品状态
            static_sample_total['R{}'.format(static_sample_total_max_row + sample_count)] = sample.monitor_task  # 监测项目
            static_sample_total['S{}'.format(static_sample_total_max_row + sample_count)] = sample.sample_count  # 样品数量
            static_sample_total['T{}'.format(static_sample_total_max_row + sample_count)] = sample.people  # 采样人

        static_excel.save(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))

        """20190417增加总体的小区监测统计表"""
        static_excel_total.save(os.path.join(output_path, '小区监测统计表.xlsx'))

        # 流量
        info_flow_sheet = info_excel['流量']
        flow_sheet_max_row = info_flow_sheet.max_row
        flows = monitor.flow.all().order_by('flow_date', 'flow_time')
        day_avg_flow_lst = []  # 平均日流量列表
        final_monitor_date = None
        flow_excel = openpyxl.load_workbook(os.path.join(output_path_for_excel_per_monitor, '流量表.xlsx'))
        flow_sheet = flow_excel['流量']
        # print(monitor.name)
        for flow_count, flow in enumerate(flows, 1):
            info_flow_sheet['A{}'.format(flow_sheet_max_row + flow_count)] = flow.id
            info_flow_sheet['B{}'.format(flow_sheet_max_row + flow_count)] = flow.monitor_point.id
            info_flow_sheet['C{}'.format(flow_sheet_max_row + flow_count)] = flow.people
            info_flow_sheet['D{}'.format(flow_sheet_max_row + flow_count)] = flow.flow_date
            info_flow_sheet['E{}'.format(flow_sheet_max_row + flow_count)] = flow.flow_time
            info_flow_sheet['F{}'.format(flow_sheet_max_row + flow_count)] = flow.time1
            info_flow_sheet['G{}'.format(flow_sheet_max_row + flow_count)] = flow.volume1
            info_flow_sheet['H{}'.format(flow_sheet_max_row + flow_count)] = flow.time2
            info_flow_sheet['I{}'.format(flow_sheet_max_row + flow_count)] = flow.volume2
            info_flow_sheet['J{}'.format(flow_sheet_max_row + flow_count)] = flow.time3
            info_flow_sheet['K{}'.format(flow_sheet_max_row + flow_count)] = flow.volume3
            info_flow_sheet['L{}'.format(flow_sheet_max_row + flow_count)] = flow.diameter
            info_flow_sheet['M{}'.format(flow_sheet_max_row + flow_count)] = flow.canal_width
            info_flow_sheet['N{}'.format(flow_sheet_max_row + flow_count)] = flow.mud_depth
            if monitor.work_function == 1:
                info_flow_sheet['O{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_lequid_level1
                info_flow_sheet['P{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_instantaneous_flow_rate1
                info_flow_sheet['Q{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_lequid_level2
                info_flow_sheet['R{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_instantaneous_flow_rate2
                info_flow_sheet['S{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_lequid_level3
                info_flow_sheet['T{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_instantaneous_flow_rate3
            elif monitor.work_function == 2:
                info_flow_sheet['O{}'.format(flow_sheet_max_row + flow_count)] = flow.square_lequid_level1
                info_flow_sheet['P{}'.format(flow_sheet_max_row + flow_count)] = flow.square_instantaneous_flow_rate1
                info_flow_sheet['Q{}'.format(flow_sheet_max_row + flow_count)] = flow.square_lequid_level2
                info_flow_sheet['R{}'.format(flow_sheet_max_row + flow_count)] = flow.square_instantaneous_flow_rate2
                info_flow_sheet['S{}'.format(flow_sheet_max_row + flow_count)] = flow.square_lequid_level3
                info_flow_sheet['T{}'.format(flow_sheet_max_row + flow_count)] = flow.square_instantaneous_flow_rate3
            # print(flow_count)
            if flow_count <= 3:
                flow_sheet['A4'] = monitor.name
                flow_sheet['B{}'.format(3 + flow_count * 3 - 2)] = flow.flow_time
                if flow_count == 1:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(4, 7):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 4]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 4]
                elif flow_count == 2:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(7, 10):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 7]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 7]
                else:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(10, 13):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 10]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 10]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A2'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表.xlsx'))
                final_monitor_date = flow.flow_date
            if 4 <= flow_count <= 6:
                flow_sheet['A19'] = monitor.name
                flow_sheet['B{}'.format(18 + (flow_count - 3) * 3 - 2)] = flow.flow_time
                if flow_count == 4:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(19, 22):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 19]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 19]
                elif flow_count == 5:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(22, 25):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 22]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 22]
                else:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(25, 28):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 25]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 25]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A17'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表.xlsx'))
                final_monitor_date = flow.flow_date
            if 7 <= flow_count <= 9:
                flow_sheet['A34'] = monitor.name
                flow_sheet['B{}'.format(18 + (flow_count - 6) * 3 - 2)] = flow.flow_time
                if flow_count == 7:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(34, 37):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 34]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 34]
                elif flow_count == 8:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(37, 40):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 37]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 37]
                else:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(40, 43):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 40]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 40]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A32'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表.xlsx'))
                final_monitor_date = flow.flow_date
            if 10 <= flow_count <= 12:
                flow_sheet['A49'] = monitor.name
                flow_sheet['B{}'.format(18 + (flow_count - 9) * 3 - 2)] = flow.flow_time
                if flow_count == 10:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(49, 52):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 49]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 49]
                elif flow_count == 11:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(52, 55):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 52]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 52]
                else:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(55, 58):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 55]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 55]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A47'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表.xlsx'))
                final_monitor_date = flow.flow_date
            if 13 <= flow_count <= 15:
                flow_sheet['A64'] = monitor.name
                flow_sheet['B{}'.format(18 + (flow_count - 12) * 3 - 2)] = flow.flow_time
                if flow_count == 13:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(64, 67):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 64]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 64]
                elif flow_count == 14:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(67, 70):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 67]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 67]
                else:
                    time_lst = [flow.time1, flow.time2, flow.time3]
                    flow_lst = [flow.volume1, flow.volume2, flow.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(70, 73):
                        temp_time = 'D' + str(j)
                        flow_sheet[temp_time].value = time_lst[j - 70]
                        temp_water = 'E' + str(j)
                        flow_sheet[temp_water].value = flow_lst[j - 70]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A62'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表.xlsx'))
                final_monitor_date = flow.flow_date
        flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表.xlsx'))
        day_avg_flow = mean(day_avg_flow_lst)  # 平均一天的流量
        flow_for_whole_day = day_avg_flow * 86400 / 1000
        static_excel = openpyxl.load_workbook(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))
        static_flow = static_excel['流量']
        static_flow['A4'] = monitor_count
        static_flow['B4'] = monitor.name
        static_flow['E4'] = monitor.geophysical_point
        static_flow['L4'] = monitor.start_time
        static_flow['M4'] = final_monitor_date
        static_flow['N4'] = '当日00:00至24:00'
        static_flow['O4'] = flow_for_whole_day
        static_flow['Q4'] = '容器法'
        static_flow['R4'] = monitor.people
        static_excel.save(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))
        info_excel.save(os.path.join(output_path_per_monitor, '监测点信息.xlsx'))

        """20190417增加总体的小区监测统计表"""
        static_excel_total = openpyxl.load_workbook(os.path.join(output_path, '小区监测统计表.xlsx'))
        static_flow_total = static_excel_total['流量']
        static_flow_total_max_row = static_flow_total.max_row
        static_flow_total['A{}'.format(static_flow_total_max_row + 1)] = monitor_count
        static_flow_total['B{}'.format(static_flow_total_max_row + 1)] = monitor.name
        static_flow_total['E{}'.format(static_flow_total_max_row + 1)] = monitor.geophysical_point
        static_flow_total['L{}'.format(static_flow_total_max_row + 1)] = monitor.start_time
        static_flow_total['M{}'.format(static_flow_total_max_row + 1)] = final_monitor_date
        static_flow_total['N{}'.format(static_flow_total_max_row + 1)] = '当日00:00至24:00'
        static_flow_total['O{}'.format(static_flow_total_max_row + 1)] = flow_for_whole_day
        static_flow_total['Q{}'.format(static_flow_total_max_row + 1)] = '容器法'
        static_flow_total['R{}'.format(static_flow_total_max_row + 1)] = monitor.people
        static_excel_total.save(os.path.join(output_path, '小区监测统计表.xlsx'))

    file_news = time_str + '.zip'
    file_news = my_setting.export_folder + os.sep + file_news
    with zipfile.ZipFile(file_news, 'w', zipfile.ZIP_DEFLATED) as zf:
        for dirpath, dirnames, filenames in os.walk(output_path):
            fpath = dirpath.replace(output_path, '')
            fpath = fpath and fpath + os.sep or ''
            for filename in filenames:
                zf.write(os.path.join(dirpath, filename), fpath + filename)
    shutil.rmtree(output_path)
    return time_str + '.zip'


def ex_circle(monitor_objs):
    """
    导出采用流速法（圆管）的监测点数据
    :param monitor_objs: 监测点对象集
    :return: time_str + '.zip': 压缩包名
    """

    time_str = time.time()
    time_str = str(time_str)  # 时间戳作为压缩包和临时文件夹名
    output_path = os.path.join(my_setting.export_folder, time_str)
    shutil.rmtree(my_setting.export_folder)
    if not os.path.exists(my_setting.export_folder):
        os.mkdir(my_setting.export_folder)
    if not os.path.exists(output_path):
        os.mkdir(output_path)
    for monitor_count, monitor in enumerate(monitor_objs, 1):
        output_path_per_monitor = os.path.join(output_path, monitor.name)  # 存放每个监测点相应文件的临时路径
        output_path_for_photo_per_monitor = os.path.join(output_path_per_monitor, '照片')  # 存放每个监测点相应照片的临时路径
        output_path_for_excel_per_monitor = os.path.join(output_path_per_monitor, '成果')  # 存放每个监测点相应Excel文件的临时路径
        if not os.path.exists(output_path_per_monitor):
            os.mkdir(output_path_per_monitor)
        if not os.path.exists(output_path_for_photo_per_monitor):
            os.mkdir(output_path_for_photo_per_monitor)
        if not os.path.exists(output_path_for_excel_per_monitor):
            os.mkdir(output_path_for_excel_per_monitor)
        shutil.copy(os.path.join(my_setting.excel_folder, '小区监测统计表.xlsx'),
                    output_path_for_excel_per_monitor)  # 监测点对应统计表excel表格
        shutil.copy(os.path.join(my_setting.excel_folder, '流量表(圆管).xlsx'),
                    output_path_for_excel_per_monitor)  # 监测点对应流量表excel表格
        shutil.copy(os.path.join(my_setting.excel_folder, '监测点信息.xlsx'), output_path_per_monitor)  # 监测点对应监测点信息excel表格
        # 监测点信息表
        info_excel = openpyxl.load_workbook(os.path.join(output_path_per_monitor, '监测点信息.xlsx'))
        monitor_sheet = info_excel['监测点']
        monitor_sheet['A{}'.format(2)] = monitor.id
        monitor_sheet['B{}'.format(2)] = monitor.name
        monitor_sheet['C{}'.format(2)] = monitor.geophysical_point
        monitor_sheet['D{}'.format(2)] = monitor.people
        if monitor.work_function == 0:
            monitor_sheet['E{}'.format(2)] = '容器法'
        elif monitor.work_function == 1:
            monitor_sheet['E{}'.format(2)] = '流速法（圆管）'
        elif monitor.work_function == 2:
            monitor_sheet['E{}'.format(2)] = '流速法（方渠）'
        elif monitor.work_function == 3:
            monitor_sheet['E{}'.format(2)] = '仪器法'
        else:
            monitor_sheet['E{}'.format(2)] = '无法监测'
        monitor_sheet['F{}'.format(2)] = monitor.not_monitor_reason
        # 外景照
        exterior_photo_lst = json.loads(monitor.exterior_photo)
        exterior_photo_lst = list_split(exterior_photo_lst)
        for exterior_photo in exterior_photo_lst:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, exterior_photo), output_path_for_photo_per_monitor)
        # 水流照
        water_flow_photo_lst = json.loads(monitor.water_flow_photo)
        water_flow_photo_lst = list_split(water_flow_photo_lst)
        for water_photo in water_flow_photo_lst:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, water_photo), output_path_for_photo_per_monitor)
        # 工作照
        work_photo_lst = json.loads(monitor.work_photo)
        work_photo_lst = list_split(work_photo_lst)
        for work_photo in work_photo_lst:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, work_photo), output_path_for_photo_per_monitor)

        # 样品
        sample_sheet = info_excel['样品']
        sample_sheet_max_row = sample_sheet.max_row
        samples = monitor.sample.all().order_by('sample_date', 'sample_time')
        static_excel = openpyxl.load_workbook(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))
        static_sample = static_excel['水质']
        for sample_count, sample in enumerate(samples, 1):
            sample_sheet['A{}'.format(sample_sheet_max_row + sample_count)] = sample.id
            sample_sheet['B{}'.format(sample_sheet_max_row + sample_count)] = sample.monitor_point.id
            sample_sheet['C{}'.format(sample_sheet_max_row + sample_count)] = sample.people
            sample_sheet['D{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_date
            sample_sheet['E{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_time
            sample_sheet['F{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_number
            sample_sheet['G{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_color
            sample_sheet['H{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_odor
            sample_sheet['I{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_turbidity
            sample_sheet['J{}'.format(sample_sheet_max_row + sample_count)] = sample.monitor_task
            sample_sheet['K{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_count
            sample_sheet['L{}'.format(sample_sheet_max_row + sample_count)] = sample.SS
            sample_sheet['M{}'.format(sample_sheet_max_row + sample_count)] = sample.NH3_N
            sample_sheet['N{}'.format(sample_sheet_max_row + sample_count)] = sample.TP
            sample_sheet['O{}'.format(sample_sheet_max_row + sample_count)] = sample.TN
            sample_sheet['P{}'.format(sample_sheet_max_row + sample_count)] = sample.COD
            sample_sheet['Q{}'.format(sample_sheet_max_row + sample_count)] = sample.BOD
            sample_sheet['R{}'.format(sample_sheet_max_row + sample_count)] = sample.AIS
            sample_sheet['S{}'.format(sample_sheet_max_row + sample_count)] = sample.AFVO
            sample_sheet['T{}'.format(sample_sheet_max_row + sample_count)] = sample.DO
            sample_sheet['U{}'.format(sample_sheet_max_row + sample_count)] = sample.FLOW
            sample_sheet['V{}'.format(sample_sheet_max_row + sample_count)] = sample.CR
            sample_sheet['W{}'.format(sample_sheet_max_row + sample_count)] = sample.ORP
            sample_sheet['X{}'.format(sample_sheet_max_row + sample_count)] = sample.SinkableS
            sample_sheet['Y{}'.format(sample_sheet_max_row + sample_count)] = sample.Sulfide
            sample_sheet['Z{}'.format(sample_sheet_max_row + sample_count)] = sample.Cyanide
            # 样品照
            sample_photo_lst = json.loads(sample.sample_photo)
            sample_photo_lst = list_split(sample_photo_lst)
            for sample_photo in sample_photo_lst:
                shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, sample_photo), output_path_for_photo_per_monitor)
            # 打开样品统计表
            static_sample['A{}'.format(4 + sample_count)] = sample_count  # 序号
            static_sample['B{}'.format(4 + sample_count)] = monitor.name  # 采样点位
            static_sample['E{}'.format(4 + sample_count)] = monitor.geophysical_point  # 物探点号
            static_sample['H{}'.format(4 + sample_count)] = sample.sample_date  # 采样日期
            static_sample['I{}'.format(4 + sample_count)] = sample.sample_time  # 采样时间段
            static_sample['J{}'.format(4 + sample_count)] = sample.sample_number  # 样品编号
            static_sample[
                'Q{}'.format(
                    4 + sample_count)] = sample.sample_color + '、' + sample.sample_odor + '、' + sample.sample_turbidity  # 样品状态
            static_sample['R{}'.format(4 + sample_count)] = sample.monitor_task  # 监测项目
            static_sample['S{}'.format(4 + sample_count)] = sample.sample_count  # 样品数量
            static_sample['T{}'.format(4 + sample_count)] = sample.people  # 采样人
        static_excel.save(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))

        # 流量
        info_flow_sheet = info_excel['流量']
        flow_sheet_max_row = info_flow_sheet.max_row
        flows = monitor.flow.all().order_by('flow_date', 'flow_time')
        day_avg_flow_lst = []  # 平均日流量列表
        final_monitor_date = None
        flow_excel = openpyxl.load_workbook(os.path.join(output_path_for_excel_per_monitor, '流量表(圆管).xlsx'))
        flow_sheet = flow_excel['流量']
        for flow_count, flow in enumerate(flows, 1):
            info_flow_sheet['A{}'.format(flow_sheet_max_row + flow_count)] = flow.id
            info_flow_sheet['B{}'.format(flow_sheet_max_row + flow_count)] = flow.monitor_point.id
            info_flow_sheet['C{}'.format(flow_sheet_max_row + flow_count)] = flow.people
            info_flow_sheet['D{}'.format(flow_sheet_max_row + flow_count)] = flow.flow_date
            info_flow_sheet['E{}'.format(flow_sheet_max_row + flow_count)] = flow.flow_time
            info_flow_sheet['F{}'.format(flow_sheet_max_row + flow_count)] = flow.time1
            info_flow_sheet['G{}'.format(flow_sheet_max_row + flow_count)] = flow.volume1
            info_flow_sheet['H{}'.format(flow_sheet_max_row + flow_count)] = flow.time2
            info_flow_sheet['I{}'.format(flow_sheet_max_row + flow_count)] = flow.volume2
            info_flow_sheet['J{}'.format(flow_sheet_max_row + flow_count)] = flow.time3
            info_flow_sheet['K{}'.format(flow_sheet_max_row + flow_count)] = flow.volume3
            info_flow_sheet['L{}'.format(flow_sheet_max_row + flow_count)] = flow.diameter
            info_flow_sheet['M{}'.format(flow_sheet_max_row + flow_count)] = flow.canal_width
            info_flow_sheet['N{}'.format(flow_sheet_max_row + flow_count)] = flow.mud_depth
            if monitor.work_function == 1:
                info_flow_sheet['O{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_lequid_level1
                info_flow_sheet['P{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_instantaneous_flow_rate1
                info_flow_sheet['Q{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_lequid_level2
                info_flow_sheet['R{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_instantaneous_flow_rate2
                info_flow_sheet['S{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_lequid_level3
                info_flow_sheet['T{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_instantaneous_flow_rate3
            elif monitor.work_function == 2:
                info_flow_sheet['O{}'.format(flow_sheet_max_row + flow_count)] = flow.square_lequid_level1
                info_flow_sheet['P{}'.format(flow_sheet_max_row + flow_count)] = flow.square_instantaneous_flow_rate1
                info_flow_sheet['Q{}'.format(flow_sheet_max_row + flow_count)] = flow.square_lequid_level2
                info_flow_sheet['R{}'.format(flow_sheet_max_row + flow_count)] = flow.square_instantaneous_flow_rate2
                info_flow_sheet['S{}'.format(flow_sheet_max_row + flow_count)] = flow.square_lequid_level3
                info_flow_sheet['T{}'.format(flow_sheet_max_row + flow_count)] = flow.square_instantaneous_flow_rate3
            # print(flow_count)
            if flow_count <= 3:
                flow_sheet['A4'] = monitor.name
                flow_sheet['B{}'.format(3 + flow_count * 3 - 2)] = flow.flow_time
                if flow_count == 1:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(4, 7):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 4]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 4]
                elif flow_count == 2:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(7, 10):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 7]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 7]
                else:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(10, 13):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 10]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 10]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A2'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表(圆管).xlsx'))
                final_monitor_date = flow.flow_date
            if 4 <= flow_count <= 6:
                flow_sheet['A19'] = monitor.name
                flow_sheet['B{}'.format(18 + (flow_count - 3) * 3 - 2)] = flow.flow_time
                if flow_count == 4:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(19, 22):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 19]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 19]
                elif flow_count == 5:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(22, 25):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 22]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 22]
                else:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(25, 28):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 25]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 25]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A17'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表(圆管).xlsx'))
                final_monitor_date = flow.flow_date
            if 7 <= flow_count <= 9:
                flow_sheet['A34'] = monitor.name
                flow_sheet['B{}'.format(18 + (flow_count - 6) * 3 - 2)] = flow.flow_time
                if flow_count == 7:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(34, 37):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 34]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 34]
                elif flow_count == 8:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(37, 40):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 37]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 37]
                else:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(40, 43):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 40]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 40]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A32'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表(圆管).xlsx'))
                final_monitor_date = flow.flow_date
            if 10 <= flow_count <= 12:
                flow_sheet['A49'] = monitor.name
                flow_sheet['B{}'.format(18 + (flow_count - 9) * 3 - 2)] = flow.flow_time
                if flow_count == 10:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(49, 52):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 49]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 49]
                elif flow_count == 11:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(52, 55):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 52]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 52]
                else:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(55, 58):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 55]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 55]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A47'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表(圆管).xlsx'))
                final_monitor_date = flow.flow_date
            if 13 <= flow_count <= 15:
                flow_sheet['A64'] = monitor.name
                flow_sheet['B{}'.format(18 + (flow_count - 12) * 3 - 2)] = flow.flow_time
                if flow_count == 13:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(64, 67):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 64]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 64]
                elif flow_count == 14:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(67, 70):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 67]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 67]
                else:
                    diameter = flow.diameter
                    cicle_lequid_lst = [flow.cicle_lequid_level1, flow.cicle_lequid_level2, flow.cicle_lequid_level3]
                    cicle_instantaneous_flow_lst = [flow.cicle_instantaneous_flow_rate1,
                                                    flow.cicle_instantaneous_flow_rate2,
                                                    flow.cicle_instantaneous_flow_rate3]
                    for j in range(70, 73):
                        flow_sheet['D{}'.format(j)].value = diameter
                        cicle_lequid = 'E' + str(j)
                        flow_sheet[cicle_lequid].value = cicle_lequid_lst[j - 70]
                        cicle_instantaneous_flow = 'F' + str(j)
                        flow_sheet[cicle_instantaneous_flow].value = cicle_instantaneous_flow_lst[j - 70]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A62'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表(圆管).xlsx'))
                final_monitor_date = flow.flow_date
        flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表(圆管).xlsx'))
        static_excel = openpyxl.load_workbook(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))
        static_flow = static_excel['流量']
        static_flow['A4'] = monitor_count
        static_flow['B4'] = monitor.name
        static_flow['E4'] = monitor.geophysical_point
        static_flow['L4'] = monitor.start_time
        static_flow['M4'] = final_monitor_date
        static_flow['N4'] = '当日00:00至24:00'
        static_flow['Q4'] = '流速法（圆管）'
        static_flow['R4'] = monitor.people
        static_excel.save(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))
        info_excel.save(os.path.join(output_path_per_monitor, '监测点信息.xlsx'))
    file_news = time_str + '.zip'
    file_news = my_setting.export_folder + os.sep + file_news
    with zipfile.ZipFile(file_news, 'w', zipfile.ZIP_DEFLATED) as zf:
        for dirpath, dirnames, filenames in os.walk(output_path):
            fpath = dirpath.replace(output_path, '')
            fpath = fpath and fpath + os.sep or ''
            for filename in filenames:
                zf.write(os.path.join(dirpath, filename), fpath + filename)
    shutil.rmtree(output_path)
    return time_str + '.zip'


def ex_square(monitor_objs):
    """
    导出采用流速法（方渠）的监测点数据
    :param monitor_objs: 监测点对象集
    :return: time_str + '.zip': 压缩包名
    """

    time_str = time.time()
    time_str = str(time_str)  # 时间戳作为压缩包和临时文件夹名
    output_path = os.path.join(my_setting.export_folder, time_str)
    shutil.rmtree(my_setting.export_folder)
    if not os.path.exists(my_setting.export_folder):
        os.mkdir(my_setting.export_folder)
    if not os.path.exists(output_path):
        os.mkdir(output_path)
    for monitor_count, monitor in enumerate(monitor_objs, 1):
        output_path_per_monitor = os.path.join(output_path, monitor.name)  # 存放每个监测点相应文件的临时路径
        output_path_for_photo_per_monitor = os.path.join(output_path_per_monitor, '照片')  # 存放每个监测点相应照片的临时路径
        output_path_for_excel_per_monitor = os.path.join(output_path_per_monitor, '成果')  # 存放每个监测点相应Excel文件的临时路径
        if not os.path.exists(output_path_per_monitor):
            os.mkdir(output_path_per_monitor)
        if not os.path.exists(output_path_for_photo_per_monitor):
            os.mkdir(output_path_for_photo_per_monitor)
        if not os.path.exists(output_path_for_excel_per_monitor):
            os.mkdir(output_path_for_excel_per_monitor)
        shutil.copy(os.path.join(my_setting.excel_folder, '小区监测统计表.xlsx'),
                    output_path_for_excel_per_monitor)  # 监测点对应统计表excel表格
        shutil.copy(os.path.join(my_setting.excel_folder, '流量表(方渠).xlsx'),
                    output_path_for_excel_per_monitor)  # 监测点对应流量表excel表格
        shutil.copy(os.path.join(my_setting.excel_folder, '监测点信息.xlsx'), output_path_per_monitor)  # 监测点对应监测点信息excel表格
        # 监测点信息表
        info_excel = openpyxl.load_workbook(os.path.join(output_path_for_excel_per_monitor, '监测点信息.xlsx'))
        monitor_sheet = info_excel['监测点']
        monitor_sheet['A{}'.format(2)] = monitor.id
        monitor_sheet['B{}'.format(2)] = monitor.name
        monitor_sheet['C{}'.format(2)] = monitor.geophysical_point
        monitor_sheet['D{}'.format(2)] = monitor.people
        if monitor.work_function == 0:
            monitor_sheet['E{}'.format(2)] = '容器法'
        elif monitor.work_function == 1:
            monitor_sheet['E{}'.format(2)] = '流速法（圆管）'
        elif monitor.work_function == 2:
            monitor_sheet['E{}'.format(2)] = '流速法（方渠）'
        elif monitor.work_function == 3:
            monitor_sheet['E{}'.format(2)] = '仪器法'
        else:
            monitor_sheet['E{}'.format(2)] = '无法监测'
        monitor_sheet['F{}'.format(2)] = monitor.not_monitor_reason
        # 外景照
        exterior_photo_lst = json.loads(monitor.exterior_photo)
        exterior_photo_lst = list_split(exterior_photo_lst)
        for exterior_photo in exterior_photo_lst:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, exterior_photo), output_path_for_photo_per_monitor)
        # 水流照
        water_flow_photo_lst = json.loads(monitor.water_flow_photo)
        water_flow_photo_lst = list_split(water_flow_photo_lst)
        for water_photo in water_flow_photo_lst:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, water_photo), output_path_for_photo_per_monitor)
        # 工作照
        work_photo_lst = json.loads(monitor.work_photo)
        work_photo_lst = list_split(work_photo_lst)
        for work_photo in work_photo_lst:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, work_photo), output_path_for_photo_per_monitor)

        # 样品
        sample_sheet = info_excel['样品']
        sample_sheet_max_row = sample_sheet.max_row
        samples = monitor.sample.all().order_by('sample_date', 'sample_time')
        static_excel = openpyxl.load_workbook(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))
        static_sample = static_excel['水质']
        for sample_count, sample in enumerate(samples, 1):
            sample_sheet['A{}'.format(sample_sheet_max_row + sample_count)] = sample.id
            sample_sheet['B{}'.format(sample_sheet_max_row + sample_count)] = sample.monitor_point.id
            sample_sheet['C{}'.format(sample_sheet_max_row + sample_count)] = sample.people
            sample_sheet['D{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_date
            sample_sheet['E{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_time
            sample_sheet['F{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_number
            sample_sheet['G{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_color
            sample_sheet['H{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_odor
            sample_sheet['I{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_turbidity
            sample_sheet['J{}'.format(sample_sheet_max_row + sample_count)] = sample.monitor_task
            sample_sheet['K{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_count
            sample_sheet['L{}'.format(sample_sheet_max_row + sample_count)] = sample.SS
            sample_sheet['M{}'.format(sample_sheet_max_row + sample_count)] = sample.NH3_N
            sample_sheet['N{}'.format(sample_sheet_max_row + sample_count)] = sample.TP
            sample_sheet['O{}'.format(sample_sheet_max_row + sample_count)] = sample.TN
            sample_sheet['P{}'.format(sample_sheet_max_row + sample_count)] = sample.COD
            sample_sheet['Q{}'.format(sample_sheet_max_row + sample_count)] = sample.BOD
            sample_sheet['R{}'.format(sample_sheet_max_row + sample_count)] = sample.AIS
            sample_sheet['S{}'.format(sample_sheet_max_row + sample_count)] = sample.AFVO
            sample_sheet['T{}'.format(sample_sheet_max_row + sample_count)] = sample.DO
            sample_sheet['U{}'.format(sample_sheet_max_row + sample_count)] = sample.FLOW
            sample_sheet['V{}'.format(sample_sheet_max_row + sample_count)] = sample.CR
            sample_sheet['W{}'.format(sample_sheet_max_row + sample_count)] = sample.ORP
            sample_sheet['X{}'.format(sample_sheet_max_row + sample_count)] = sample.SinkableS
            sample_sheet['Y{}'.format(sample_sheet_max_row + sample_count)] = sample.Sulfide
            sample_sheet['Z{}'.format(sample_sheet_max_row + sample_count)] = sample.Cyanide
            # 样品照
            sample_photo_lst = json.loads(sample.sample_photo)
            sample_photo_lst = list_split(sample_photo_lst)
            for sample_photo in sample_photo_lst:
                shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, sample_photo), output_path_for_photo_per_monitor)
            # 打开样品统计表
            static_sample['A{}'.format(4 + sample_count)] = sample_count  # 序号
            static_sample['B{}'.format(4 + sample_count)] = monitor.name  # 采样点位
            static_sample['E{}'.format(4 + sample_count)] = monitor.geophysical_point  # 物探点号
            static_sample['H{}'.format(4 + sample_count)] = sample.sample_date  # 采样日期
            static_sample['I{}'.format(4 + sample_count)] = sample.sample_time  # 采样时间段
            static_sample['J{}'.format(4 + sample_count)] = sample.sample_number  # 样品编号
            static_sample[
                'Q{}'.format(
                    4 + sample_count)] = sample.sample_color + '、' + sample.sample_odor + '、' + sample.sample_turbidity  # 样品状态
            static_sample['R{}'.format(4 + sample_count)] = sample.monitor_task  # 监测项目
            static_sample['S{}'.format(4 + sample_count)] = sample.sample_count  # 样品数量
            static_sample['T{}'.format(4 + sample_count)] = sample.people  # 采样人
        static_excel.save(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))

        # 流量
        info_flow_sheet = info_excel['流量']
        flow_sheet_max_row = info_flow_sheet.max_row
        flows = monitor.flow.all().order_by('flow_date', 'flow_time')
        day_avg_flow_lst = []  # 平均日流量列表
        final_monitor_date = None
        flow_excel = openpyxl.load_workbook(os.path.join(output_path_for_excel_per_monitor, '流量表(方渠).xlsx'))
        flow_sheet = flow_excel['流量']
        for flow_count, flow in enumerate(flows, 1):
            info_flow_sheet['A{}'.format(flow_sheet_max_row + flow_count)] = flow.id
            info_flow_sheet['B{}'.format(flow_sheet_max_row + flow_count)] = flow.monitor_point.id
            info_flow_sheet['C{}'.format(flow_sheet_max_row + flow_count)] = flow.people
            info_flow_sheet['D{}'.format(flow_sheet_max_row + flow_count)] = flow.flow_date
            info_flow_sheet['E{}'.format(flow_sheet_max_row + flow_count)] = flow.flow_time
            info_flow_sheet['F{}'.format(flow_sheet_max_row + flow_count)] = flow.time1
            info_flow_sheet['G{}'.format(flow_sheet_max_row + flow_count)] = flow.volume1
            info_flow_sheet['H{}'.format(flow_sheet_max_row + flow_count)] = flow.time2
            info_flow_sheet['I{}'.format(flow_sheet_max_row + flow_count)] = flow.volume2
            info_flow_sheet['J{}'.format(flow_sheet_max_row + flow_count)] = flow.time3
            info_flow_sheet['K{}'.format(flow_sheet_max_row + flow_count)] = flow.volume3
            info_flow_sheet['L{}'.format(flow_sheet_max_row + flow_count)] = flow.diameter
            info_flow_sheet['M{}'.format(flow_sheet_max_row + flow_count)] = flow.canal_width
            info_flow_sheet['N{}'.format(flow_sheet_max_row + flow_count)] = flow.mud_depth
            if monitor.work_function == 1:
                info_flow_sheet['O{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_lequid_level1
                info_flow_sheet['P{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_instantaneous_flow_rate1
                info_flow_sheet['Q{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_lequid_level2
                info_flow_sheet['R{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_instantaneous_flow_rate2
                info_flow_sheet['S{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_lequid_level3
                info_flow_sheet['T{}'.format(flow_sheet_max_row + flow_count)] = flow.cicle_instantaneous_flow_rate3
            elif monitor.work_function == 2:
                info_flow_sheet['O{}'.format(flow_sheet_max_row + flow_count)] = flow.square_lequid_level1
                info_flow_sheet['P{}'.format(flow_sheet_max_row + flow_count)] = flow.square_instantaneous_flow_rate1
                info_flow_sheet['Q{}'.format(flow_sheet_max_row + flow_count)] = flow.square_lequid_level2
                info_flow_sheet['R{}'.format(flow_sheet_max_row + flow_count)] = flow.square_instantaneous_flow_rate2
                info_flow_sheet['S{}'.format(flow_sheet_max_row + flow_count)] = flow.square_lequid_level3
                info_flow_sheet['T{}'.format(flow_sheet_max_row + flow_count)] = flow.square_instantaneous_flow_rate3
            if flow_count <= 3:
                flow_sheet['A4'] = monitor.name
                flow_sheet['B{}'.format(3 + flow_count * 3 - 2)] = flow.flow_time
                if flow_count == 1:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(4, 7):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 4]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 4]
                elif flow_count == 2:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(7, 10):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 7]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 7]
                else:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(10, 13):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 10]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 10]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A2'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表(方渠).xlsx'))
                final_monitor_date = flow.flow_date
            if 4 <= flow_count <= 6:
                flow_sheet['A19'] = monitor.name
                flow_sheet['B{}'.format(18 + (flow_count - 3) * 3 - 2)] = flow.flow_time
                if flow_count == 4:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(19, 22):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 19]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 19]
                elif flow_count == 5:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(22, 25):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 22]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 22]
                else:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(25, 28):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 25]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 25]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A17'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表(方渠).xlsx'))
                final_monitor_date = flow.flow_date
            if 7 <= flow_count <= 9:
                flow_sheet['A34'] = monitor.name
                flow_sheet['B{}'.format(18 + (flow_count - 6) * 3 - 2)] = flow.flow_time
                if flow_count == 7:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(34, 37):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 34]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 34]
                elif flow_count == 8:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(37, 40):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 37]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 37]
                else:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(40, 43):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 40]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 40]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A32'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表(方渠).xlsx'))
                final_monitor_date = flow.flow_date
            if 10 <= flow_count <= 12:
                flow_sheet['A49'] = monitor.name
                flow_sheet['B{}'.format(18 + (flow_count - 9) * 3 - 2)] = flow.flow_time
                if flow_count == 10:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(49, 52):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 49]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 49]
                elif flow_count == 11:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(52, 55):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 52]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 52]
                else:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(55, 58):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 55]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 55]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A47'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表(方渠).xlsx'))
                final_monitor_date = flow.flow_date
            if 13 <= flow_count <= 15:
                flow_sheet['A64'] = monitor.name
                flow_sheet['B{}'.format(18 + (flow_count - 12) * 3 - 2)] = flow.flow_time
                if flow_count == 13:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(64, 67):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 64]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 64]
                elif flow_count == 14:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(67, 70):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 67]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 67]
                else:
                    canal_width = flow.canal_width
                    square_lequid_lst = [flow.square_lequid_level1, flow.square_lequid_level2,
                                         flow.square_lequid_level3]
                    square_instantaneous_flow_lst = [flow.square_instantaneous_flow_rate1,
                                                     flow.square_instantaneous_flow_rate2,
                                                     flow.square_instantaneous_flow_rate3]
                    flow1 = canal_width * flow.square_lequid_level1 * flow.square_instantaneous_flow_rate1
                    flow2 = canal_width * flow.square_lequid_level2 * flow.square_instantaneous_flow_rate2
                    flow3 = canal_width * flow.square_lequid_level3 * flow.square_instantaneous_flow_rate3
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(70, 73):
                        flow_sheet['D{}'.format(j)].value = canal_width
                        square_lequid = 'E' + str(j)
                        flow_sheet[square_lequid].value = square_lequid_lst[j - 70]
                        square_instantaneous_flow = 'F' + str(j)
                        flow_sheet[square_instantaneous_flow].value = square_instantaneous_flow_lst[j - 70]
                intro = f'监测日期：{flow.flow_date}                    观测者：{monitor.people}                    检查者：'
                flow_sheet['A62'] = intro
                flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表(方渠).xlsx'))
                final_monitor_date = flow.flow_date
        flow_excel.save(os.path.join(output_path_for_excel_per_monitor, '流量表(方渠).xlsx'))
        day_avg_flow = mean(day_avg_flow_lst)  # 平均一天的流量
        flow_for_whole_day = day_avg_flow * 86400
        static_excel = openpyxl.load_workbook(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))
        static_flow = static_excel['流量']
        static_flow['A4'] = monitor_count
        static_flow['B4'] = monitor.name
        static_flow['E4'] = monitor.geophysical_point
        static_flow['L4'] = monitor.start_time
        static_flow['M4'] = final_monitor_date
        static_flow['N4'] = '当日00:00至24:00'
        static_flow['O4'] = flow_for_whole_day
        static_flow['Q4'] = '流速法（方渠）'
        static_flow['R4'] = monitor.people
        static_excel.save(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))
        info_excel.save(os.path.join(output_path_per_monitor, '监测点信息.xlsx'))
    file_news = time_str + '.zip'
    file_news = my_setting.export_folder + os.sep + file_news
    with zipfile.ZipFile(file_news, 'w', zipfile.ZIP_DEFLATED) as zf:
        for dirpath, dirnames, filenames in os.walk(output_path):
            fpath = dirpath.replace(output_path, '')
            fpath = fpath and fpath + os.sep or ''
            for filename in filenames:
                zf.write(os.path.join(dirpath, filename), fpath + filename)
    shutil.rmtree(output_path)
    return time_str + '.zip'


def ex_machine(monitor_objs):
    """
    导出采用仪器法的监测点数据
    :param monitor_objs: 监测点对象集
    :return: time_str + '.zip': 压缩包名
    """

    time_str = time.time()
    time_str = str(time_str)  # 时间戳作为压缩包和临时文件夹名
    output_path = os.path.join(my_setting.export_folder, time_str)
    shutil.rmtree(my_setting.export_folder)
    if not os.path.exists(my_setting.export_folder):
        os.mkdir(my_setting.export_folder)
    if not os.path.exists(output_path):
        os.mkdir(output_path)

    for monitor_count, monitor in enumerate(monitor_objs, 1):
        output_path_per_monitor = os.path.join(output_path, monitor.name)  # 存放每个监测点相应文件的临时路径
        output_path_for_photo_per_monitor = os.path.join(output_path_per_monitor, '照片')  # 存放每个监测点相应照片的临时路径
        output_path_for_excel_per_monitor = os.path.join(output_path_per_monitor, '成果')  # 存放每个监测点相应Excel文件的临时路径
        if not os.path.exists(output_path_per_monitor):
            os.mkdir(output_path_per_monitor)
        if not os.path.exists(output_path_for_photo_per_monitor):
            os.mkdir(output_path_for_photo_per_monitor)
        if not os.path.exists(output_path_for_excel_per_monitor):
            os.mkdir(output_path_for_excel_per_monitor)
        shutil.copy(os.path.join(my_setting.excel_folder, '小区监测统计表.xlsx'),
                    output_path_for_excel_per_monitor)  # 监测点对应统计表excel表格
        shutil.copy(os.path.join(my_setting.excel_folder, '监测点信息.xlsx'),
                    output_path_for_excel_per_monitor)  # 监测点对应监测点信息excel表格
        # 监测点信息表
        info_excel = openpyxl.load_workbook(os.path.join(output_path_per_monitor, '监测点信息.xlsx'))
        monitor_sheet = info_excel['监测点']
        monitor_sheet['A{}'.format(2)] = monitor.id
        monitor_sheet['B{}'.format(2)] = monitor.name
        monitor_sheet['C{}'.format(2)] = monitor.geophysical_point
        monitor_sheet['D{}'.format(2)] = monitor.people
        if monitor.work_function == 0:
            monitor_sheet['E{}'.format(2)] = '容器法'
        elif monitor.work_function == 1:
            monitor_sheet['E{}'.format(2)] = '流速法（圆管）'
        elif monitor.work_function == 2:
            monitor_sheet['E{}'.format(2)] = '流速法（方渠）'
        elif monitor.work_function == 3:
            monitor_sheet['E{}'.format(2)] = '仪器法'
        else:
            monitor_sheet['E{}'.format(2)] = '无法监测'
        monitor_sheet['F{}'.format(2)] = monitor.not_monitor_reason
        # 现状照
        status_photo = json.loads(monitor.status_photo)
        status_photo = list_split(status_photo)
        for sp in status_photo:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, sp), output_path_for_photo_per_monitor)

        probe_photo = json.loads(monitor.probe_photo)
        probe_photo = list_split(probe_photo)
        for pp in probe_photo:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, pp), output_path_for_photo_per_monitor)

        machine_photo = json.loads(monitor.machine_photo)
        machine_photo = list_split(machine_photo)
        for mp in machine_photo:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, mp), output_path_for_photo_per_monitor)

        setup_photo = json.loads(monitor.setup_photo)
        setup_photo = list_split(setup_photo)
        for sep in setup_photo:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, sep), output_path_for_photo_per_monitor)

        work_photo = json.loads(monitor.work_photo)
        work_photo = list_split(work_photo)
        for wp in work_photo:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, wp), output_path_for_photo_per_monitor)

        # 样品
        sample_sheet = info_excel['样品']
        sample_sheet_max_row = sample_sheet.max_row
        samples = monitor.sample.all().order_by('sample_date', 'sample_time')
        static_excel = openpyxl.load_workbook(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))
        static_sample = static_excel['水质']
        for sample_count, sample in enumerate(samples, 1):
            sample_sheet['A{}'.format(sample_sheet_max_row + sample_count)] = sample.id
            sample_sheet['B{}'.format(sample_sheet_max_row + sample_count)] = sample.monitor_point.id
            sample_sheet['C{}'.format(sample_sheet_max_row + sample_count)] = sample.people
            sample_sheet['D{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_date
            sample_sheet['E{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_time
            sample_sheet['F{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_number
            sample_sheet['G{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_color
            sample_sheet['H{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_odor
            sample_sheet['I{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_turbidity
            sample_sheet['J{}'.format(sample_sheet_max_row + sample_count)] = sample.monitor_task
            sample_sheet['K{}'.format(sample_sheet_max_row + sample_count)] = sample.sample_count
            sample_sheet['L{}'.format(sample_sheet_max_row + sample_count)] = sample.SS
            sample_sheet['M{}'.format(sample_sheet_max_row + sample_count)] = sample.NH3_N
            sample_sheet['N{}'.format(sample_sheet_max_row + sample_count)] = sample.TP
            sample_sheet['O{}'.format(sample_sheet_max_row + sample_count)] = sample.TN
            sample_sheet['P{}'.format(sample_sheet_max_row + sample_count)] = sample.COD
            sample_sheet['Q{}'.format(sample_sheet_max_row + sample_count)] = sample.BOD
            sample_sheet['R{}'.format(sample_sheet_max_row + sample_count)] = sample.AIS
            sample_sheet['S{}'.format(sample_sheet_max_row + sample_count)] = sample.AFVO
            sample_sheet['T{}'.format(sample_sheet_max_row + sample_count)] = sample.DO
            sample_sheet['U{}'.format(sample_sheet_max_row + sample_count)] = sample.FLOW
            sample_sheet['V{}'.format(sample_sheet_max_row + sample_count)] = sample.CR
            sample_sheet['W{}'.format(sample_sheet_max_row + sample_count)] = sample.ORP
            sample_sheet['X{}'.format(sample_sheet_max_row + sample_count)] = sample.SinkableS
            sample_sheet['Y{}'.format(sample_sheet_max_row + sample_count)] = sample.Sulfide
            sample_sheet['Z{}'.format(sample_sheet_max_row + sample_count)] = sample.Cyanide
            # 样品照
            sample_photo_lst = json.loads(sample.sample_photo)
            sample_photo_lst = list_split(sample_photo_lst)
            for sample_photo in sample_photo_lst:
                shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, sample_photo), output_path_for_photo_per_monitor)
            # 打开样品统计表
            static_sample['A{}'.format(4 + sample_count)] = sample_count  # 序号
            static_sample['B{}'.format(4 + sample_count)] = monitor.name  # 采样点位
            static_sample['E{}'.format(4 + sample_count)] = monitor.geophysical_point  # 物探点号
            static_sample['H{}'.format(4 + sample_count)] = sample.sample_date  # 采样日期
            static_sample['I{}'.format(4 + sample_count)] = sample.sample_time  # 采样时间段
            static_sample['J{}'.format(4 + sample_count)] = sample.sample_number  # 样品编号
            static_sample[
                'Q{}'.format(
                    4 + sample_count)] = sample.sample_color + '、' + sample.sample_odor + '、' + sample.sample_turbidity  # 样品状态
            static_sample['R{}'.format(4 + sample_count)] = sample.monitor_task  # 监测项目
            static_sample['S{}'.format(4 + sample_count)] = sample.sample_count  # 样品数量
            static_sample['T{}'.format(4 + sample_count)] = sample.people  # 采样人
        static_excel.save(os.path.join(output_path_for_excel_per_monitor, '小区监测统计表.xlsx'))
        info_excel.save(os.path.join(output_path_per_monitor, '监测点信息.xlsx'))
    file_news = time_str + '.zip'
    file_news = my_setting.export_folder + os.sep + file_news
    with zipfile.ZipFile(file_news, 'w', zipfile.ZIP_DEFLATED) as zf:
        for dirpath, dirnames, filenames in os.walk(output_path):
            fpath = dirpath.replace(output_path, '')
            fpath = fpath and fpath + os.sep or ''
            for filename in filenames:
                zf.write(os.path.join(dirpath, filename), fpath + filename)
    shutil.rmtree(output_path)
    return time_str + '.zip'


def ex_unable(monitor_objs):
    """
     导出采用仪器法的监测点数据
     :param monitor_objs: 监测点对象集
     :return: time_str + '.zip': 压缩包名
     """

    time_str = time.time()
    time_str = str(time_str)  # 时间戳作为压缩包和临时文件夹名
    output_path = os.path.join(my_setting.export_folder, time_str)
    shutil.rmtree(my_setting.export_folder)
    if not os.path.exists(my_setting.export_folder):
        os.mkdir(my_setting.export_folder)
    if not os.path.exists(output_path):
        os.mkdir(output_path)

    for monitor_count, monitor in enumerate(monitor_objs, 1):
        output_path_per_monitor = os.path.join(output_path, monitor.name)  # 存放每个监测点相应文件的临时路径
        output_path_for_photo_per_monitor = os.path.join(output_path_per_monitor, '照片')  # 存放每个监测点相应照片的临时路径
        if not os.path.exists(output_path_per_monitor):
            os.mkdir(output_path_per_monitor)
        if not os.path.exists(output_path_for_photo_per_monitor):
            os.mkdir(output_path_for_photo_per_monitor)
        shutil.copy(os.path.join(my_setting.excel_folder, '小区监测统计表.xlsx'), output_path_per_monitor)  # 监测点对应统计表excel表格
        shutil.copy(os.path.join(my_setting.excel_folder, '监测点信息.xlsx'), output_path_per_monitor)  # 监测点对应监测点信息excel表格

        # 监测点信息表
        info_excel = openpyxl.load_workbook(os.path.join(output_path_per_monitor, '监测点信息.xlsx'))
        monitor_sheet = info_excel['监测点']
        monitor_sheet['A{}'.format(2)] = monitor.id
        monitor_sheet['B{}'.format(2)] = monitor.name
        monitor_sheet['C{}'.format(2)] = monitor.geophysical_point
        monitor_sheet['D{}'.format(2)] = monitor.people
        if monitor.work_function == 0:
            monitor_sheet['E{}'.format(2)] = '容器法'
        elif monitor.work_function == 1:
            monitor_sheet['E{}'.format(2)] = '流速法（圆管）'
        elif monitor.work_function == 2:
            monitor_sheet['E{}'.format(2)] = '流速法（方渠）'
        elif monitor.work_function == 3:
            monitor_sheet['E{}'.format(2)] = '仪器法'
        else:
            monitor_sheet['E{}'.format(2)] = '无法监测'
        monitor_sheet['F{}'.format(2)] = monitor.not_monitor_reason

        exterior_photo = json.loads(monitor.exterior_photo)
        exterior_photo = list_split(exterior_photo)
        for ep in exterior_photo:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, ep), output_path_per_monitor)

        water_flow_photo = json.loads(monitor.water_flow_photo)
        water_flow_photo = list_split(water_flow_photo)
        for wfp in water_flow_photo:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, wfp), output_path_per_monitor)

        static_excel = openpyxl.load_workbook(os.path.join(output_path_per_monitor, '小区监测统计表.xlsx'))
        static_flow = static_excel['流量']
        static_flow['A4'] = monitor_count
        static_flow['B4'] = monitor.name
        static_flow['E4'] = monitor.geophysical_point
        static_flow['L4'] = monitor.start_time
        static_flow['N4'] = '当日00:00至24:00'
        static_flow['P4'] = monitor.not_monitor_reason
        static_flow['Q4'] = '无法检测'
        static_flow['R4'] = monitor.people
        static_excel.save(os.path.join(output_path_per_monitor, '小区监测统计表.xlsx'))
        info_excel.save(os.path.join(output_path_per_monitor, '监测点信息.xlsx'))
    file_news = time_str + '.zip'
    file_news = my_setting.export_folder + os.sep + file_news
    with zipfile.ZipFile(file_news, 'w', zipfile.ZIP_DEFLATED) as zf:
        for dirpath, dirnames, filenames in os.walk(output_path):
            fpath = dirpath.replace(output_path, '')
            fpath = fpath and fpath + os.sep or ''
            for filename in filenames:
                zf.write(os.path.join(dirpath, filename), fpath + filename)
    shutil.rmtree(output_path)
    return time_str + '.zip'
